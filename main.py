# ----------------------------
# Importation dyal les libraries
# ----------------------------
from datetime import datetime
import math
import matplotlib.pyplot as plt
import pandas as pd
from fpdf import FPDF
import os
import sys

# ----------------------------
# Installation automatique dyal les dépendances
# ----------------------------
try:
    import openpyxl
except ImportError:
    print("🔧 Installation dyal openpyxl...")
    os.system(f'{sys.executable} -m pip install openpyxl')
    import openpyxl

try:
    import seaborn
except ImportError:
    print("🔧 Installation dyal seaborn...")
    os.system(f'{sys.executable} -m pip install seaborn')
    import seaborn

# ----------------------------
# Configuration initiale
# ----------------------------
plt.style.use('ggplot')  # Style alternatif si seaborn marche pas

# ----------------------------
# Fonctions d'interface utilisateur
# ----------------------------
def get_user_input():
    """Recueille les inputs utilisateur avec validation"""
    while True:
        try:
            print("\n" + "="*50)
            print("📊 ADS CALCULATOR PRO 2025")
            print("="*50 + "\n")
            
            user_budget = float(input("📌 Dkhel Budget dyalek (DH): "))
            if user_budget <= 0:
                print("❌ Budget khass ykun akbar mn 0!")
                continue
                
            target_age = input("🎯 Target Age (Ex: 18-35): ")
            target_location = input("📍 Target Location (Ex: Casablanca, Maroc): ")
            
            ad_duration = int(input("⏳ Duration dyal Campagne (jours): "))
            if ad_duration <= 0:
                print("❌ Duration khass tkun akbar mn 0!")
                continue
                
            business_type = input("🏢 Type dyal Business (E-commerce/Service/Education): ").lower()
            
            return user_budget, target_age, target_location, ad_duration, business_type
            
        except ValueError:
            print("❌ Dkhel nombre valid!")

# ----------------------------
# Fonctions de génération de rapports
# ----------------------------
def generate_chart(daily_budget, clicks, impressions, duration):
    """Génère un graphique de performance"""
    try:
        plt.figure(figsize=(10, 5))
        days = list(range(1, duration + 1))
        
        # Données cumulatives
        cumulative_clicks = [clicks/duration * day for day in days]
        cumulative_impressions = [impressions/duration * day for day in days]
        
        plt.plot(days, cumulative_clicks, label='Clicks', marker='o', color='blue')
        plt.plot(days, cumulative_impressions, label='Impressions', marker='s', color='green')
        
        plt.title("Performance de Campagne (Cumulatif)")
        plt.xlabel("Jours")
        plt.ylabel("Nombre")
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        
        plt.savefig('ads_performance.png', dpi=300)
        print("✅ Graphique mnsab f 'ads_performance.png'")
        plt.close()
    except Exception as e:
        print(f"⚠️ Erreur graphique: {str(e)}")

def export_to_excel(data, filename='ads_report.xlsx'):
    """Exporte les données vers Excel"""
    try:
        df = pd.DataFrame(data)
        
        # Création d'un writer Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Rapport')
            
            # Accès au workbook et worksheet pour le formatage
            workbook = writer.book
            worksheet = writer.sheets['Rapport']
            
            # Formatage de base
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2
                
            print(f"✅ Rapport Excel mnsab f '{filename}'")
    except Exception as e:
        print(f"⚠️ Erreur Excel: {str(e)}")

def generate_pdf_report(results, filename='ads_report.pdf'):
    """Génère un rapport PDF professionnel"""
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        # En-tête
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, "Rapport de Campagne Publicitaire", 0, 1, 'C')
        pdf.ln(5)
        
        # Informations de base
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, "Paramètres de Campagne:", 0, 1)
        pdf.set_font('Arial', '', 12)
        
        infos = [
            ["Budget Total", f"{results['budget']} DH"],
            ["Budget Quotidien", f"{results['daily_budget']:.2f} DH"],
            ["Durée", f"{results['duration']} jours"],
            ["Cible Age", results['target_age']],
            ["Localisation", results['target_location']],
            ["Type Business", results['business_type'].capitalize()]
        ]
        
        for info in infos:
            pdf.cell(90, 10, info[0], 1)
            pdf.cell(0, 10, info[1], 1, 1)
        
        pdf.ln(5)
        
        # Performance
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, "Performance Estimée:", 0, 1)
        pdf.set_font('Arial', '', 12)
        
        metrics = [
            ["Impressions", f"{results['impressions']:,}"],
            ["Clicks", f"{results['clicks']:,}"],
            ["Leads Potentiels", f"{results['leads']}"],
            ["CPC Moyen", f"{results['cpc']:.2f} DH"],
            ["ROAS Estimé", f"{results['roas']}%"]
        ]
        
        for metric in metrics:
            pdf.cell(90, 10, metric[0], 1)
            pdf.cell(0, 10, metric[1], 1, 1)
        
        # Ajout du graphique si disponible
        if os.path.exists('ads_performance.png'):
            pdf.ln(10)
            pdf.image('ads_performance.png', x=10, w=180)
        
        pdf.output(filename)
        print(f"✅ Rapport PDF mnsab f '{filename}'")
    except Exception as e:
        print(f"⚠️ Erreur PDF: {str(e)}")

# ----------------------------
# Fonction de calcul principale
# ----------------------------
def calculate_ads_performance():
    """Fonction principale qui orchestre tous les calculs"""
    # Récupération des inputs
    budget, age, location, duration, business_type = get_user_input()
    
    # Coefficients par type de business
    coefficients = {
        'e-commerce': {'click': 3.2, 'cpm': 120, 'roas': 4.0},
        'service': {'click': 1.8, 'cpm': 200, 'roas': 2.5},
        'education': {'click': 2.1, 'cpm': 180, 'roas': 3.2}
    }
    
    # Application des coefficients
    coeff = coefficients.get(business_type, {'click': 2.5, 'cpm': 150, 'roas': 3.0})
    
    # Calculs
    daily_budget = budget / duration
    clicks = math.floor(budget * coeff['click'])
    impressions = math.floor(budget * coeff['cpm'])
    cpc = round(budget / clicks, 2) if clicks > 0 else 0
    leads = math.floor(clicks * 0.15)
    roas = round(budget * coeff['roas'] / 1000, 2)
    
    # Affichage des résultats
    print("\n" + "="*50)
    print("📈 RESULTATS D'ANALYSE")
    print("="*50)
    print(f"• Budget/jour: {daily_budget:.2f} DH")
    print(f"• Impressions: {impressions:,}")
    print(f"• Clicks: {clicks:,}")
    print(f"• Leads: {leads}")
    print(f"• CPC: {cpc:.2f} DH")
    print(f"• ROAS: {roas}%")
    
    # Préparation des données pour les rapports
    results = {
        'budget': budget,
        'daily_budget': daily_budget,
        'duration': duration,
        'target_age': age,
        'target_location': location,
        'business_type': business_type,
        'impressions': impressions,
        'clicks': clicks,
        'leads': leads,
        'cpc': cpc,
        'roas': roas
    }
    
    excel_data = {
        'Métrique': list(results.keys()),
        'Valeur': list(results.values())
    }
    
    # Génération des rapports
    generate_chart(daily_budget, clicks, impressions, duration)
    export_to_excel(excel_data)
    generate_pdf_report(results)
    
    # Conseils
    print("\n💡 Conseils 2025:")
    print("- Zid vidéos courtes pour plus d'engagement")
    print("- Testiw les audiences similaires (Lookalike)")
    print("- Utilise UTM pour mieux tracker")
    print("="*50 + "\n")

# ----------------------------
# Point d'entrée principal
# ----------------------------
if __name__ == "__main__":
    calculate_ads_performance()